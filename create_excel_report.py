import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KiemThu_NoSQL"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Styling definitions
font_title = Font(name="Arial", size=16, bold=True, color="1F4E79")
font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)
font_pass = Font(name="Arial", size=10, bold=True, color="006100")

fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_summary = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# 1. Header Information
ws.merge_cells("A1:K1")
ws["A1"] = "BÁO CÁO KẾT QUẢ KIỂM THỬ TÍNH NĂNG CSDL NOSQL (REDIS & NEO4J)"
ws["A1"].font = font_title
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:K2")
ws["A2"] = "Dự án: Web Đặt Vé Xem Phim (Movana Cinema) | Môn học: Các hệ CSDL NoSQL"
ws["A2"].font = font_subtitle
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

info_data = [
    ("Sinh viên thực hiện:", "1. Tăng Gia Huy   |   2. Lê Ngọc Anh (Người kiểm thử chính)   |   3. Đào Trọng Nguyên Vũ"),
    ("Môi trường kiểm thử:", "ASP.NET MVC 5 (.NET 4.5), C#, Redis Server v7.0, Neo4j Graph v5.x, MS SQL Server"),
    ("Ngày thực hiện:", "08/08/2026")
]

row_idx = 4
for label, val in info_data:
    ws.cell(row=row_idx, column=1, value=label).font = font_bold
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=11)
    ws.cell(row=row_idx, column=2, value=val).font = font_regular
    row_idx += 1

row_idx += 1 # Empty row space

# 2. Table Headers (11 columns)
headers = [
    "STT", "Phân Hệ NoSQL", "Mã Test Case", "Tên Chức Năng", 
    "Kịch Bản Kiểm Thử (Scenario)", "Các Bước Thực Hiện (Test Steps)", 
    "Dữ Liệu Đầu Vào (Input)", "Kết Quả Mong Cho (Expected)", 
    "Kết Quả Thực Tế (Actual)", "Trạng Thái", "Người Thực Hiện"
]

header_row = row_idx
for col_idx, text in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[header_row].height = 28

# 3. Test Case Data with Step-by-Step and Tester Name "Lê Ngọc Anh"
test_cases = [
    (
        1, "Redis Cache", "TC_REDIS_01", "Khóa ghế tạm thời (Atomic Seat Lock)",
        "Khách hàng chọn ghế trên sơ đồ phòng chiếu, hệ thống kích hoạt SETNX trên Redis để giữ ghế nguyên tử trong 90 giây.",
        "1. Đăng nhập tài khoản khách hàng\n2. Chọn phim & suất chiếu bất kỳ\n3. Chọn các vị trí ghế mong muốn\n4. Kiểm tra Redis key 'seatlock:{lc}:{ghe}' đếm ngược 90s",
        "LichChieuID: 10, SeatIDs: ['A1', 'A2'], Duration: 90s",
        "Redis tạo Key 'seatlock:10:A1' có TTL=90s. Người dùng khác chọn cùng ghế sẽ bị từ chối.",
        "Khóa ghế thành công trên Redis RAM. Hết 90s Redis tự giải phóng khóa chuẩn xác.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        2, "Redis Cache", "TC_REDIS_02", "Giỏ hàng thanh toán (User Cart Hash)",
        "Lưu tạm thông tin giỏ hàng vé xem phim của người dùng vào Redis Hash với TTL 10 phút (600s) trong quá trình thanh toán.",
        "1. Chọn ghế và bấm 'Xác nhận đặt vé'\n2. Chuyển sang trang Thanh toán\n3. Kiểm tra Redis Hash 'cart:{username}' chứa giỏ hàng\n4. Hủy giao dịch hoặc thanh toán xong -> Redis tự xóa giỏ hàng",
        "Username: 'leanh325', Hash: {LichChieuID, Ghe, TongTien}, TTL: 600s",
        "Tạo Key Redis Hash 'cart:leanh325' chứa thông tin đơn hàng. Xóa giỏ hàng khi hủy hoặc thanh toán xong.",
        "Thông tin giỏ hàng lưu cực nhanh trên RAM Redis, tự hủy khi hết hạn 600s hoặc mua vé xong.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        3, "Redis Cache", "TC_REDIS_03", "Xác thực OTP Gmail (Gmail SMTP + Redis TTL)",
        "Tạo mã OTP ngẫu nhiên 6 số lưu Redis đếm ngược 120s và gửi Gmail SMTP thực tế tới hộp thư khách hàng.",
        "1. Tại trang Thanh toán, bấm nút 'LẤY MÃ OTP'\n2. Hệ thống sinh OTP 6 số lưu Redis (TTL 120s) và gửi Mail thật qua Gmail SMTP\n3. Kiểm tra hộp thư Gmail nhận mã OTP\n4. Nhập mã OTP và bấm Thanh Toán",
        "Username: 'leanh325', Email: 'anh874343@gmail.com', OTP: 6-digit",
        "Mã OTP lưu trên Redis với Key 'otp:checkout:leanh325' (TTL 120s). Nhập đúng OTP mới cho thanh toán.",
        "Gmail SMTP gửi OTP thực tế về hòm thư, Redis xác thực OTP thành công trong 120s và xóa ngay sau khi dùng.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        4, "Redis Cache", "TC_REDIS_04", "Quản lý User Session trên RAM",
        "Lưu trữ phiên làm việc (Session) của người dùng đăng nhập trên RAM Redis với TTL 30 phút (1800s).",
        "1. Nhập Username và Password bấm Đăng nhập\n2. Hệ thống lưu Hash 'session:user:{username}' trên RAM Redis TTL 30 phút\n3. Điều hướng giữa các trang trong ứng dụng\n4. Bấm 'Đăng xuất' -> Redis xóa Session Key",
        "Username: 'leanh325', SessionHash: {UserID, FullName, GroupID}",
        "Key 'session:user:leanh325' lưu trên Redis. Khi người dùng bấm Sign Out, Redis xóa ngay Session Key.",
        "Phiên đăng nhập phản hồi siêu tốc, tự động hủy phiên trên Redis khi đăng xuất thành công.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        5, "Neo4j Graph", "TC_NEO4J_01", "Gợi ý Phim Thông Minh (Recommendation Engine)",
        "Truy vấn Cypher đề xuất các bộ phim CÙNG THỂ LOẠI với phim người dùng đã xem/thả tim (Không trùng lặp, Top 4 phim).",
        "1. Đăng nhập tài khoản và bấm Thả tim (❤️) hoặc Đặt vé phim 'Mùi Cỏ Cháy' (Lịch sử)\n2. Hệ thống gọi truy vấn Cypher tìm phim cùng thể loại trên Neo4j Graph\n3. Quay về Trang chủ xem mục 'GỢI Ý PHIM DÀNH CHO BẠN'\n4. Kiểm tra danh sách hiển thị Top 4 phim Lịch sử chuẩn xác",
        "Username: 'leanh325', MovieBooked: 'Mùi Cỏ Cháy' (Lịch sử)",
        "Trả về Top 4 phim cùng thể loại 'Lịch sử' (Đào Phở Piano, Mưa Đỏ, Hà Nội 12 Ngày Đêm). Bù phim hot khác nếu hết thể loại.",
        "Truy vấn Cypher trả về chuẩn xác các phim Lịch sử không bị trùng lặp, hiển thị mượt mà trên trang chủ.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        6, "Neo4j Graph", "TC_NEO4J_02", "Thống kê Bảng xếp hạng Top Phim Thịnh Hành",
        "Thống kê Top Phim Đặt Vé và Yêu Thích nhiều nhất thời gian thực dựa trên các nút quan hệ đồ thị Neo4j.",
        "1. Khách hàng thực hiện Đặt vé và Thả tim các bộ phim\n2. Đồ thị Neo4j tự động tích lũy số lượng quan hệ (:BOOKED) và (:FAVORITE)\n3. Vào Trang chủ hoặc trang Phim Đang Chiếu\n4. Kiểm tra Bảng xếp hạng Top Phim hiển thị theo thứ tự đặt vé",
        "Graph Patterns: (:User)-[:BOOKED]->(:Movie), (:User)-[:FAVORITE]->(:Movie)",
        "Trả về bảng xếp hạng Top Phim theo tổng lượt đặt vé và lượt thả tim cao nhất.",
        "Đồ thị Neo4j tính toán lượt tương tác chuẩn xác, hiển thị đầy đủ trên Bảng xếp hạng.",
        "PASSED", "Lê Ngọc Anh"
    ),
    (
        7, "Neo4j Graph", "TC_NEO4J_03", "Thả tim Yêu thích Phim (Toggle Favorite Relationship)",
        "Bật/Tắt quan hệ Yêu thích (:FAVORITE) giữa Nút Người Dùng và Nút Bộ Phim trên Đồ thị Neo4j.",
        "1. Tìm bộ phim mong muốn trên danh sách\n2. Bấm nút Trái tim (❤️) trên góc card phim\n3. Hệ thống gọi AJAX tới Neo4jController/ToggleFavorite tạo quan hệ (u)-[:FAVORITE]->(m)\n4. Kiểm tra biểu tượng trái tim đổi màu và bấm lại để hủy yêu thích",
        "Username: 'leanh325', MovieID: 42 ('Mùi Cỏ Cháy')",
        "Tạo mới quan hệ (u)-[:FAVORITE]->(m) và gắn thể loại (m)-[:BELONGS_TO]->(g) nếu chưa thả tim, xóa quan hệ nếu bấm lại.",
        "Neo4j cập nhật quan hệ đồ thị tức thì, biểu tượng trái tim đổi màu tương ứng trên giao diện.",
        "PASSED", "Lê Ngọc Anh"
    )
]

data_start_row = header_row + 1
current_row = data_start_row

for item in test_cases:
    stt, sys_name, tc_code, feature, scenario, steps, input_data, expected, actual, status, tester = item
    
    is_zebra = (current_row % 2 == 0)
    row_fill = fill_zebra if is_zebra else None

    ws.cell(row=current_row, column=1, value=stt).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=current_row, column=2, value=sys_name).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=current_row, column=3, value=tc_code).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=current_row, column=4, value=feature).alignment = Alignment(horizontal="left", vertical="top")
    ws.cell(row=current_row, column=5, value=scenario).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=current_row, column=6, value=steps).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=current_row, column=7, value=input_data).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=current_row, column=8, value=expected).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=current_row, column=9, value=actual).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    cell_status = ws.cell(row=current_row, column=10, value=status)
    cell_status.alignment = Alignment(horizontal="center", vertical="top")
    cell_status.font = font_pass
    cell_status.fill = fill_pass

    cell_tester = ws.cell(row=current_row, column=11, value=tester)
    cell_tester.alignment = Alignment(horizontal="center", vertical="top")
    cell_tester.font = font_bold

    for col in range(1, 12):
        c = ws.cell(row=current_row, column=col)
        if col != 10 and row_fill:
            c.fill = row_fill
        c.border = thin_border
        if col != 10:
            c.font = font_regular
            if col in [1, 2, 3, 11]:
                c.font = font_bold

    ws.row_dimensions[current_row].height = 65
    current_row += 1

# 4. Summary Table
current_row += 1
ws.cell(row=current_row, column=1, value="TỔNG HỢP KẾT QUẢ KIỂM THỬ").font = font_bold

summary_start = current_row + 1
sum_data = [
    ("Tổng số Test Cases thực hiện:", 7),
    ("Số lượng Test Cases PASSED:", 7),
    ("Số lượng Test Cases FAILED:", 0),
    ("Tỷ lệ thành công (Pass Rate):", "100%"),
    ("Người thực hiện kiểm thử chính:", "Lê Ngọc Anh")
]

for idx, (lbl, val) in enumerate(sum_data, start=summary_start):
    c1 = ws.cell(row=idx, column=1, value=lbl)
    c1.font = font_bold
    c1.fill = fill_summary
    c1.border = thin_border
    
    c2 = ws.cell(row=idx, column=2, value=val)
    c2.font = font_bold
    c2.fill = fill_summary
    c2.alignment = Alignment(horizontal="center")
    c2.border = thin_border

# Column width formatting
col_widths = {
    1: 6,   # STT
    2: 14,  # System
    3: 15,  # TC Code
    4: 28,  # Feature
    5: 35,  # Scenario
    6: 40,  # Test Steps
    7: 30,  # Input
    8: 35,  # Expected
    9: 35,  # Actual
    10: 14, # Status
    11: 18  # Tester
}

for col_idx, width in col_widths.items():
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

# Save to Desktop
desktop_path = r"C:\Users\Le Ngoc Anh\Desktop\KiemThu_NoSQL_Redis_Neo4j.xlsx"
wb.save(desktop_path)
print(f"Successfully updated report with Test Steps and Tester 'Lê Ngọc Anh' at: {desktop_path}")
